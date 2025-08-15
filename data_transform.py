# -*- coding: utf-8 -*-
"""
Created on Thu Jul  3 13:25:27 2025

@author: garre
"""
"""
Code to copy point-based data from a data sheet used for data input into a ICASA data template.
All columns that should be transferred need to have the same variable name as in the ICASA templete 
or a mapping must be provided (mapping_file).
The data file from where the data is copied needs to have a column 
that specifies the ODMF number assigned to the point where the sample was taken
or a mapping from your ids to the ODMF must be provided (id_file).

If you run into problems due to datatypes of columns that are empty in one or the other data sheet, 
try inserting a dummy row with values of the correct type and delete it later.
"""

"""
Provide the complete path to the input data sheet (excel format) and sheet name within the workbook.
"""
input_file = "H:/Data/LI600_2025.xlsx"
input_sheet = "all"

"""
Provide the complete path to the icasa template (excel format) and sheet name to which the data should be copied.
"""
template_file = "H:/Data/FORMULA_point_data_2.xlsx"
template_sheet = "PORO_FLUORO"

"""
Specify whether you want to provide a mapping table instead of using the ICASA variable names in your input_file.
If true, provide a mapping table that contains the ICASA varaible names in the first column. 
and your variable names in the second column.
The file can contain more variable names than used in the sheets you want to transform
"""

use_mapping =False
mapping_file = "H:/Data/test_mapping.xlsx"
mapping_sheet = "variables"

"""
Specify whether you want to provide a custom-id to treatment_number table 
instead of using the ODMF treatment numbers in your input_file.
If true, provide a mapping table that contains the treatment_number (ODMF number) in the first column 
and your ids in the second column.
The file can contain more ids than used in the sheets you want to transform.
Also provide the column name of your custom ID (string).
"""

use_custom_ids = False
id_file = "H:/Data/test_mapping.xlsx"
id_sheet = "ids"
id_name = "id"


"""
# specify whether data should be summarized over tecnical replicates (RP) on the same date_of_measurement, 
# If you choose to summarize:
#   only columns containing numbers or time can be transferred
#   a column "date_of_measurement" must be in the input data sheet, replicates can have any name
#   standard deviation will be added if the column is included in the template
# if replicates are present but no summary is intended
#    a column RP must specify replicate numbers both in the input and the template file
"""

summarize_samples = True

"""
#provide unit change information (optional)
#provide a dictionary of all variables that need a unit change and 
#the corresponding factors to tranform from the input unit to the output unit 
#(e.g.plant_height: input centimeter, output meter, provide "PHTD":0.01)
"""

unit_change = {}


"""
#specify whether you want to overwrite existing values in the template_file with values from the input file
#this can be used if wrong values have been imported before. 
#Be aware out that if some (wrong) values are stored in template, they will not be overwritten by importing empty lines.
# In this case you need to delete the value manually. This is also true for previously calcualted standard deviations.
#Choose False if no data from the template file should be lost.
"""

overwrite_values = False



import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# imporating the data, excudong the top rows from the template

input_data = pd.read_excel(input_file, sheet_name = input_sheet)

template_data = pd.read_excel(template_file, sheet_name=template_sheet, skiprows=3)

#rename input data columns and/or rows

if use_mapping:
    mapping = pd.read_excel(mapping_file, sheet_name=mapping_sheet)
    rename_dict = dict(zip(mapping.iloc[:,1],mapping.iloc[:,0]))
    input_data = input_data.rename(columns=rename_dict)
    
if use_custom_ids:
    ids = pd.read_excel(id_file, sheet_name = id_sheet)
    rename_id_dict = dict(zip(ids.iloc[:,1], ids.iloc[:,0]))
    input_data[id_name] = input_data[id_name].map(rename_id_dict)
    input_data = input_data.rename(columns={id_name: "treatment_number"})

#checking for common columns

common_cols = input_data.columns.intersection(template_data.columns)

#transforming time column if applicable, in order to be summarized

if summarize_samples and "time_of_measurement" in common_cols:
    input_data["time_of_measurement"] = pd.to_timedelta(input_data["time_of_measurement"].astype(str))

#subsetting data 

input_data_subset = input_data.loc[:,common_cols]

#transforming units

for entri in unit_change:
    input_data_subset[entri] = input_data_subset[entri]*unit_change[entri]

#summarizing data and computing standard deviation

if summarize_samples:
    input_data_subset = input_data_subset.groupby(["treatment_number", 'date_of_measurement']).agg(['mean', 'std', 'count']).reset_index() #reset index avoids merged cells for same treatment_number

    new_columns = []
    for col in input_data_subset.columns:
        if col[1] == '':  # This is a grouping column like ('treatment_number', '') or ('date_of_measurement', '')
            new_columns.append(col[0])
        elif col[1] == 'mean':
            new_columns.append(col[0])  # Keep original name
        elif col[1] == 'std':
            new_columns.append(col[0] + '_stdev')  # Append 'S' for std
        elif col[1] == 'count':
            if "number_of_samples" not in new_columns: #just include the first occurence, assuming some for all variables (no NAs for just one variable in the same datasheet)
                 new_columns.append("number_of_samples")
            else:
                new_columns.append("to_delete")
    
    input_data_subset.columns = new_columns
    
    if "to_delete" in input_data_subset.columns:
        input_data_subset = input_data_subset.drop(columns = ["to_delete"])
    

common_cols_2 = input_data_subset.columns.intersection(template_data.columns) #needed to include standard deviation if intended and find correct keys

input_data_subset = input_data_subset[common_cols_2]

# unite the template and input data

if "date_of_measurement" and "RP" in common_cols_2:
    keys = ["treatment_number", "date_of_measurement", "RP"]
elif "RP" in common_cols_2:
    keys = ["treatment_number", "RP"]
elif "date_of_measurement" in common_cols_2:
    keys = ["treatment_number", "date_of_measurement"]
else:
    keys = ["treatment_number"]

data_cols = [col for col in common_cols_2 if col not in keys]

merged_data = pd.merge(template_data, input_data_subset, on = keys, how = 'outer', suffixes = ("_t", "_i"))

if overwrite_values:
    for col in data_cols:
        merged_data[col] = merged_data[f"{col}_i"].combine_first(merged_data[f"{col}_t"]) #creates combination columns that have the original names (stored in data_cols), containing value from input, only if input has no value, use value from template
else:
    for col in data_cols:
        merged_data[col] = merged_data[f"{col}_t"].combine_first(merged_data[f"{col}_i"]) #creates combination columns that have the original names (stored in data_cols), containing value from template, only if template was no value, use value from input
    
final_data = merged_data[template_data.columns] #merged data contains combination column and the columns with indexes, keep only keys and combination columns and empty columns from template sheet

# write the new template into the old excel sheet (and format the columns)

# Load workbook and worksheet
wb = load_workbook(template_file)
ws = wb[template_sheet]

# Write new data starting at row 4
for r_idx, row in enumerate(dataframe_to_rows(final_data, index=False, header=True), start=4):
    for c_idx, value in enumerate(row, start=1):
        ws.cell(row=r_idx, column=c_idx, value=value)

# get headers
header = [cell.value for cell in ws[4]]

# date_of_measurement column formatting
if "date_of_measurement" in common_cols_2:
    date_col_idx = header.index("date_of_measurement") + 1  # 1-based indexing
    for row in ws.iter_rows(min_row=5, min_col=date_col_idx, max_col=date_col_idx):
        row[0].number_format = "yyyy-mm-dd"

# time_of_measurement column formatting
if "time_of_measurement" in common_cols_2:
    time_col_idx = header.index("time_of_measurement") + 1
    for row in ws.iter_rows(min_row=5, min_col=time_col_idx, max_col=time_col_idx):
        row[0].number_format = "hh:mm:ss"

wb.save(template_file)       