# -*- coding: utf-8 -*-
"""
Created on Mon Oct  6 16:25:49 2025
Revised: Added config file for credentials

@author: garre

This script exports data from the ODMF database and pastes it into an excel workbook template of the ICASA format provided by the user.
The user can choose to provide either the id of a valuetype for which datasets from all available sites will be exported 
or the id of a site for which datasets from all available valuetypes will the exportet. The program extracts the information about the corresponding
ICASA variables names stored in the comment with each valuetype in ODMF, exports and summarized the corresponding datasets, 
converts the units and aggregats the data to daily timesteps as specified in the comment, searched the provided template workbook 
for the sheet in which the valuetype is stored and pastes the final data into the excel sheet while merging to data previously stored in the sheet.

"""

import os
import configparser
from odmfclient import login
import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import logging

    
def data_by_valuetype(api, valuetype_id, project_id, start_date, end_date) -> pd.DataFrame: 
    """
    Exports all data from a given ODMF database and a given project that stores 
    values of the given type, between the start date and the end date (included).

    Parameters
    ----------
    api : ?
        Odmfclient login with url, username and password. Make sure that you have access to the data you want to export.
    valuetype_id : integer
        ID given in the ODMF system to the valuetype for which data should be exported.
    project_id : integer
        ID given in the ODMF system to the project from which data should be exported.
    start_date : string
        First date for which data should be exported in format yyyy-mm-dd.
    end_date : string
        Last date for which data should be exported in format yyyy-mm-dd.

    Returns
    -------
    data_total : pd.DataFrame
        extracted data sorted by site, level, and time

    """
    datasets = api.dataset.list(valuetype=valuetype_id, project=project_id)
    data_total = pd.DataFrame(columns=["time", "value", "site", "level"])
    end_time = end_date+"T23:59:59Z"
    for dataset_id in datasets:
        data = api.dataset.values_parquet(dsid=dataset_id, start=start_date, end=end_time)
        dataset_obj = api.dataset(dsid=dataset_id)
        site = dataset_obj["site"]["id"]
        data["site"]=site
        level = dataset_obj["level"]
        data["level"]=level
        data_total = pd.concat([data_total, data], ignore_index = True)
    data_total["date"]=data_total["time"].dt.normalize()
    data_total["time"]=data_total["time"] - data_total["date"]
    return data_total


def data_by_site(api, site_id, project_id, start_date, end_date) -> dict:
    '''
    Exports all data from a given site and a given project within ODMF database, between the start date and the end date (included),
    as one dataset per valuetype sorted in a dictionary.

    Parameters
    ----------
    api : ?
        Odmfclient login with url, username and password. Make sure that you have access to the data you want to export.
    site_id : integer
        ID given in the ODMF system to the site for which data should be exported.
    project_id : integer
        ID given in the ODMF system to the project from which data should be exported.
    start_date : string
        First date for which data should be exported in format yyyy-mm-dd.
    end_date : string
        Last date for which data should be exported in format yyyy-mm-dd.

    Returns
    -------
    data_dict : dict
        Dictionary of pd.DataFrames for each dataset stored for the site.

    '''
    datasets = api.dataset.list(site=site_id, project=project_id)
    end_time = end_date+"T23:59:59Z"
    data_dict = {}
    for dataset_id in datasets:
        data = api.dataset.values_parquet(dsid=dataset_id, start=start_date, end=end_time)
        dataset_obj = api.dataset(dsid=dataset_id)
        level = dataset_obj["level"]
        data["level"]=level
        valuetype_id = dataset_obj["valuetype"]["id"]
        data["date"]=data["time"].dt.normalize()
        data["time"]=data["time"] - data["date"]
        data_dict[valuetype_id] = data
    return data_dict


def agg_data_daily(df, function_name) -> pd.DataFrame:
    """
    Aggregates data exported from ODMF e.g. by data_by_valuetype per day using the given aggregation function.

    Parameters
    ----------
    df : Data.Frame
        A dataframe containing values sorted by site, level and time.
    function_name : string
        A aggregation function auch as mean, sum, min, max.

    Returns
    -------
    data_summed : pd.DataFrame
        DataFrame containing the aggregated data.

    """
    df["level"] = df["level"].fillna("None") #makes None a string to make grouping possible if there are no levels in the data
    
    data_summed = (
    df
    .groupby(
        [
            pd.Grouper(key='date', freq='D'),
            'site',
            'level'
        ]
    )                              
    .agg(value=("value", function_name),time=("time", "mean"))                                    
    .reset_index()                             
    )
    return data_summed


def extract_ICASA_info (api, valuetype_id, project_id) -> list:
    '''
    Extracts information about the ICASA variable corresponding to the given value_type.
    A list of all ICASA variables listed in the comment is returned.

    Parameters
    ----------
    api : ?
        Odmfclient login with url, username and password. Make sure that you have access to the data you want to export.
    valuetype_id : integer
        ID given in the ODMF system to tcdhe valuetype for which the information should be extracted.
    project_id : integer
        ID of a project in ODMF that uses the valuetype to ensure access to the information via datasets.

    Returns
    -------
    all_info : list of dictionaries
        dictionaries containing the ICASA variable name, the unit conversion factor 
        and the agrregation function for transform from the value_type to the ICASA variable_name for each ICASA variable_name that is given in the ODMF comment.
        
    '''
    datasets = api.dataset.list(valuetype=valuetype_id, project=project_id)
    first_dataset = datasets[1]
    first_dataset_obj = api.dataset(dsid=first_dataset)
    valuetype_info =first_dataset_obj["valuetype"]["comment"]
   
    pattern = re.compile(
    r'''
    ICASA:\s*
    (?P<Variable_name>[^*\n,]+)          # Variable_name (mandatory)
    (?:\*(?P<conversion>\d+(?:\.\d+)?))?   # *conversion (optional but required if * present)
    (?:,\s*(?P<aggregation>\S+))?          # ,aggregation (optional but required if , present)
    \s*$                        # line end
    ''',
    re.VERBOSE | re.MULTILINE
    )
   
    all_info = []
    
    for extraction in pattern.finditer(valuetype_info):
        ICASA_dict=extraction.groupdict()
        factor = ICASA_dict.get("conversion")
        ICASA_dict["conversion"]=float(factor) if factor else None
        all_info.append(ICASA_dict)
        
    return all_info


def find_ICASA_sheet_by_variable_name (variable_name, file_path) -> str:
    '''
    Searches the given ICASA template workbook for the data sheet in which the gven ICASA valiable name is listed.

    Parameters
    ----------
    variable_name : string
        Name of the ICASA variable to localize.
    file_path : string
        Path to the ICASA template to search in.

    Returns
    -------
    sheet_name: str

    '''
    wb = load_workbook(file_path, data_only=True)

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        row4_values = [cell.value for cell in ws[4]]
        if any(str(value) == str(variable_name) for value in row4_values if value is not None):
            sheet_name = sheet
    
    if "sheet_name" not in locals():
        raise ValueError("Variable name is not found in the provided ICASA template (check for spaces!)")
    
    return sheet_name
            

def merge_new_data_to_ICASA (new_data, template_data, site_col= "sampling_location_number", date_col = "date_of_measurement", time_col = "time_of_measurement", level_col = None, overwrite=False) -> pd.DataFrame:
    '''
    Merges data provided in the format as returned by the export functions (columns site, date, time, value and level)
    into an ICASA template sheet. 

    Parameters
    ----------
    new_data : dataframe
        Containing the data to add to the template.
    template_data : dataframe
        Containing the sheet of the template to which the new data should be added.
    site_col: string, optional
        Name of the column in the ICASA template into which ODMF site information should be pasted (e.g. weather_station_id). The default is sampling_location_number.
    date_col: string, optional
        Name of the column in the ICASA template into which ODMF date information should be pasted (e.g. weather_date). The default is date_of_measurement.
    time_col: string, optional
        Name of the column in the ICASA template into which time split from ODMF data information should be pasted. The default is time_of_measurement.
    level_col: string, optional
        Name of the column in the ICASA template into which ODMF layer information should be pasted (e.g. me_soil_layer_bot_depth)
    overwrite : boolean, optional
        Switch to allow overwriting existing values in the ICASA template with new data. The default is False.

    Returns
    -------
    final_data: pd.DataFrame
        templated_data into which common variables of new_data were merged.

    '''
    common_cols = new_data.columns.intersection(template_data.columns)

    new_data_subset = new_data.loc[:,common_cols]
    
    candidate_keys = [site_col, date_col, time_col, level_col]

    keys = [k for k in candidate_keys if k in common_cols]
    
    data_cols = [col for col in common_cols if col not in keys]

    merged_data = pd.merge(template_data, new_data_subset, on = keys, how = 'outer', suffixes = ("_t", "_i"))

    if overwrite:
        for col in data_cols:
            merged_data[col] = merged_data[f"{col}_i"].combine_first(merged_data[f"{col}_t"]) #creates combination columns that have the original names (stored in data_cols), containing value from new_data. Only if new_data has no value, use value from template_data.
    else:
        for col in data_cols:
            merged_data[col] = merged_data[f"{col}_t"].combine_first(merged_data[f"{col}_i"]) #creates combination columns that have the original names (stored in data_cols), containing value from template_data. Only if template_data has no value, use value from new_data.

    final_data = merged_data[template_data.columns] #drop colums that where created while merging and not needed after combining
    
    return final_data


def write_combined_data_to_excel (combined_data, file_path, sheet_name, date_col = "date_of_measurement", time_col = "time_of_measurement"):
    '''
    Writes data in the format the ICASA template (as returned by merge_data_to_ICASA) 
    to the given ICASA template Excel file while keeping the rest of the workbook unchanged.
    
    Parameters
    ----------
    combined_data : dataframe
        Dataframe containing data from ODMF merged with the existing template data.
    file_path: string
        Path to the ICASA template file into which the data should be pasted. This file will be partially overwritten so store a copy elsewhere to not risk of loosing data or the original template!
    sheet_name : string
        Name of the sheet within the ICASA template file into which the data should be pasted.
    date_col: string, optional
        Name of the column in the ICASA template into which ODMF date information should be pasted (e.g. weather_date). The default is date_of_measurement.
    time_col: string, optional
        Name of the column in the ICASA template into which time split from ODMF data information should be pasted. The default is time_of_measurement.

    Returns
    -------
    None.

    '''
    wb = load_workbook(file_path)
    ws = wb[sheet_name]

    # Write new data starting at row 4
    for r_idx, row in enumerate(dataframe_to_rows(combined_data, index=False, header=True), start=4):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    # get headers
    header = [cell.value for cell in ws[4]]

    # date_of_measurement column formatting
    if date_col in header:
        date_col_idx = header.index(date_col) + 1  # 1-based indexing
        for row in ws.iter_rows(min_row=5, min_col=date_col_idx, max_col=date_col_idx):
            row[0].number_format = "yyyy-mm-dd"

    # time_of_measurement column formatting
    if time_col in header:
        time_col_idx = header.index(time_col) + 1
        for row in ws.iter_rows(min_row=5, min_col=time_col_idx, max_col=time_col_idx):
            row[0].number_format = "hh:mm:ss"

    wb.save(file_path) 


def data_to_ICASA_by_valuetype (api, valuetype_id, project_id, start_date, end_date, file_path, site_col= "sampling_location_number", date_col = "date_of_measurement", time_col = "time_of_measurement",  level_col = None, overwrite =False):
    '''
    Extracts data from the ODMF system for the given valuetype and project (all sites), 
    converts it to the format of the given ICASA template and writes into the given ICASA file.

    Parameters
    ----------
    api : ?
        Odmfclient login with url, username and password. Make sure that you have access to the data you want to export.
    valuetype_id : integer
        ID given in the ODMF system to the valuetype for which data should be exported.
    project_id : integer
        ID given in the ODMF system to the project from which data should be exported.
        DESCRIPTION.
    start_date : string
        First date for which data should be exported in format yyyy-mm-dd.
    end_date : String
        Last date for which data should be exported in format yyyy-mm-dd.
    file_path: string
        Path to the ICASA template file into which the data should be pasted. This file will be partially overwritten so store a copy elsewhere to not risk of loosing data or the original template!
    site_col: string, optional
        Name of the column in the ICASA template into which ODMF site information should be pasted (e.g. weather_station_id). The default is sampling_location_number.
    date_col: string, optional
        Name of the column in the ICASA template into which ODMF date information should be pasted (e.g. weather_date). The default is date_of_measurement.
    time_col: string, optional
        Name of the column in the ICASA template into which time split from ODMF data information should be pasted. The default is time_of_measurement.
    level_col: string, optional
        Name of the column in the ICASA template into which ODMF layer information should be pasted (e.g. me_soil_layer_bot_depth). The default is None.
    overwrite: boolean, optional
        Switch to allow overwriting existing values in the ICASA template with new data. The default is False.
    Returns
    -------
    None.

    '''
    all_ICASA_infos = extract_ICASA_info(api, valuetype_id, project_id)
    
    for ICASA_info in all_ICASA_infos:
        ICASA_name = ICASA_info["Variable_name"]
        ICASA_conversion = ICASA_info["conversion"]
        ICASA_aggregation = ICASA_info["aggregation"]
        
        data = data_by_valuetype(api, valuetype_id, project_id, start_date, end_date)
        
        if data.empty:
            logging.warning(f"No dataset could be exported for {ICASA_name}. Check whether (1) Datasets are present for the given site and you have access to them via the project and api provided, (2) the datsets have entries in the time span you provided, and (3) you are connected to a network that gives you access to ODMF.")
            continue
        
        if ICASA_conversion != None:
           data["value"] = data["value"]/ICASA_conversion
        
        if ICASA_aggregation != None:
            data = agg_data_daily(data, ICASA_aggregation)
         
        data = data.rename(columns={"date": date_col, "time": time_col, "site": site_col, "level": level_col, "value": ICASA_name})
            
        try:
            ICASA_sheet_name = find_ICASA_sheet_by_variable_name(ICASA_name, file_path)
        except:
            logging.waring(f"No sheet with the variable {ICASA_name} could be found in the template. Skipped {ICASA_name}")
            continue
        
        template_data = pd.read_excel(file_path, sheet_name=ICASA_sheet_name, skiprows=3)
        
        try:
            template_data[date_col]=pd.to_datetime(template_data[date_col])
        except:
            logging.warning(f"there is no {date_col} in the same sheet as {ICASA_name}. Skipped {ICASA_name}")
            continue
        
        if time_col in template_data.columns:
            template_data[time_col]=pd.to_timedelta(template_data[time_col])
        
        combined_data = merge_new_data_to_ICASA(data, template_data, site_col, date_col, time_col, level_col, overwrite)
            
        write_combined_data_to_excel(combined_data, file_path, ICASA_sheet_name, date_col, time_col)


def data_to_ICASA_by_site (api, site_id, project_id, start_date, end_date, file_path, site_col= "weather_station_id", date_col = "date_of_measurement", time_col = "time_of_measurement",  level_col = None, overwrite =False):
    '''
    Extracts data from the ODMF system for the given site and project (all valuetypes), 
    converts it to the format of the given ICASA template and writes into the given ICASA file.

    Parameters
    ----------
    api : ?
        Odmfclient login with url, username and password. Make sure that you have access to the data you want to export.
    site_id : integer
        ID given in the ODMF system to the site for which data should be exported.
    project_id : integer
        ID given in the ODMF system to the project from which data should be exported.
        DESCRIPTION.
    start_date : string
        First date for which data should be exported in format yyyy-mm-dd.
    end_date : String
        Last date for which data should be exported in format yyyy-mm-dd.
    file_path: string
        Path to the ICASA template file into which the data should be pasted. This file will be partially overwritten so store a copy elsewhere to not risk of loosing data or the original template!
    site_col: string, optional
        Name of the column in the ICASA template into which ODMF site information should be pasted. The default is weather_station_id.
    date_col: string, optional
        Name of the column in the ICASA template into which ODMF date information should be pasted (e.g. weather_date). The default is date_of_measurement.
    time_col: string, optional
        Name of the column in the ICASA template into which time split from ODMF data information should be pasted. The default is time_of_measurement.
    level_col: string, optional
        Name of the column in the ICASA template into which ODMF layer information should be pasted (e.g. me_soil_layer_bot_depth). The default is None.
    overwrite: boolean, optional
        Switch to allow overwriting existing values in the ICASA template with new data. The default is False.
    Returns
    -------
    None.

    '''
    data_dict = data_by_site(api, site_id, project_id, start_date, end_date)
    
    for valuetype_id in data_dict:
        all_ICASA_infos = extract_ICASA_info(api, valuetype_id, project_id)
        
        for ICASA_info in all_ICASA_infos:
            ICASA_name = ICASA_info["Variable_name"]
            ICASA_conversion = ICASA_info["conversion"]
            ICASA_aggregation = ICASA_info["aggregation"]
                
            data = data_dict[valuetype_id]
            data["site"] = site_id
                
            if data.empty:
                logging.warning(f"No dataset could be exported for {ICASA_name}. Check whether (1) Datasets are present for the given site and you have access to them via the project and api provided, (2) the datsets have entries in the time span you provided, and (3) you are connected to a network that gives you access to ODMF.")
                continue
                
            if ICASA_conversion != None:
                data["value"] = data["value"]/ICASA_conversion
                
            if ICASA_aggregation != None:
                data = agg_data_daily(data, ICASA_aggregation)
                 
            data = data.rename(columns={"date": date_col, "time": time_col, "site": site_col, "level": level_col, "value": ICASA_name})
                    
            try:
                ICASA_sheet_name = find_ICASA_sheet_by_variable_name(ICASA_name, file_path)
            except:
                logging.warning(f"No sheet with the variable {ICASA_name} could be found in the template. Skipped {ICASA_name}")
                continue
                
            template_data = pd.read_excel(file_path, sheet_name=ICASA_sheet_name, skiprows=3)
                
            try:
                template_data[date_col]=pd.to_datetime(template_data[date_col])
            except:
                logging.warning(f"there is no {date_col} in the same sheet as {ICASA_name}. Skipped {ICASA_name}")
                continue
            
            if time_col in template_data.columns:
                template_data[time_col]=pd.to_timedelta(template_data[time_col])
                
            combined_data = merge_new_data_to_ICASA(data, template_data, site_col, date_col, time_col, level_col, overwrite)
                    
            write_combined_data_to_excel(combined_data, file_path, ICASA_sheet_name, date_col, time_col)   


if __name__ == "__main__":
    
    config_path = "config.ini"
    config = configparser.ConfigParser()
    config.read(config_path)

    url = config["odmf"]["url"]
    username = config["odmf"]["username"]
    password = config["odmf"]["password"]

    template_file = "ICASA_for_agroforstry_input_test.xlsx"
    BASE_DIR = os.path.abspath(os.path.dirname(__file__)) #do not run this line alone, only works when entire scrip is run
    input_path = os.path.join(BASE_DIR, template_file)
    
    with login(url, username, password) as api:
        
        # FORMULA project id: 7
    
        #ICASA_test_output = data_to_ICASA_by_valuetype(api, valuetype_id=10, project_id=7, start_date="2025-10_18", end_date="2025-10-20", file_path=template_file, level_col = "me_soil_layer_top_depth")
        ICASA_weather_test_output = data_to_ICASA_by_site(api, site_id=3817, project_id=None, start_date="2026-01-03", end_date="2026-01-06", file_path=template_file, date_col = "weather_date")