# -*- coding: utf-8 -*-
"""
Created on Mon Oct 20 12:05:35 2025

@author: garre

The Script produces a glossary from an excel sheet containing many ICASA data sheets (according to the template issued August 2025), 
listing the two rows for "Variable_Name" and "Unit_or_type", by default rows 3 and 4 of each sheet by sheet-names.
An seperate excel file can be produced, from where the glossary can be copiedcorresponding ICASA data sheet.
Rows of the resulting glossary that originate from sheets not containing data (e.g. ReadMe) need to be deleted later by hand.
The glossary can be "enriched" with two more rows of Variable information (e.g. Code_Query and Description) 
from a previous glossary or the ICASA Dictionary.

"""

import pandas as pd
from openpyxl import load_workbook
import os


def print_sheet_names (src_path: str) -> None:
    wb = load_workbook(src_path, read_only=True, data_only=True)
    ordered_sheets = wb.sheetnames
    for name in ordered_sheets:
        print(name)
    wb.close()
    
def extract_two_rows(df: pd.DataFrame) -> pd.DataFrame:
    """
    Helper: keep only rows 3 and 4 (index 2 & 3) and drop columns that are empty
    in **both** rows. Return a DataFrame of shape (n_filled_cells, 2)
    where column 0 = row‑3 value, column 1 = row‑4 value.
    """
    rows = df.iloc[[2, 3]]

    transposed = rows.T.reset_index(drop=True)
    transposed.columns = ["Unit_or_type", "Variable_Name"]          
    return transposed


def build_glossary_dataframe(
    src_path: str,
    row_indices: tuple[int, int] = (2, 3)
) -> pd.DataFrame:
    """
    Reads the given excel file and extracts the given columns (index starts at  0) from each sheet.

    The output has three columns:
        - Row3 : value from Excel row 3
        - Row4 : value from Excel row 4
        - Sheet: name of the sheet the pair came from
    """
    sheet_dict = pd.read_excel(
            src_path,
            sheet_name=None,
            header=None,
            dtype=str,
            engine="openpyxl"
        )
    
    all_blocks = []                         

    for sheet_name, raw_df in sheet_dict.items(): #looping through the dictionary and keeping both the keys (assigned to sheet_name) and the values (in this case dataframes assigned to raw_df)
        
        sliced_df = raw_df.iloc[:max(row_indices)+1]   # safe cut‑off to spare memory

        block = extract_two_rows(sliced_df)
        
        block["Sheet"] = sheet_name          
        # Keep the column order that the final CSV expects
        block = block[["Sheet", "Variable_Name", "Unit_or_type",]]
        all_blocks.append(block)

    glossary_df = pd.concat(all_blocks, axis=0, ignore_index=True)

    return glossary_df


def write_glossary_to_new_file(
    glossary_df: pd.DataFrame,
    dest_path: str,
    sheet_name: str = "glossary"
) -> None:
    """
    Writes *glossary_df* to a brand‑new workbook *dest_path*.
    If the file already exists it will be overwritten.
    """
    # ExcelWriter with engine=openpyxl creates a fresh file when mode='w'
    with pd.ExcelWriter(dest_path, engine="openpyxl", mode="w") as writer:
        glossary_df.to_excel(writer, sheet_name=sheet_name, index=False)
    print(f"[DONE] Glossary saved to '{dest_path}' (sheet name: '{sheet_name}').")
    
def enrich_glossary_with_metadata(
    glossary_df: pd.DataFrame,
    src_path: str,
    glossary_sheet_name: str = "Glossary",
    var_col: str = "Variable_Name",
    col1: str = "Code_Query",
    col2: str = "Description",
) -> pd.DataFrame:
    """
    Append two rows (e.g. Code_Query & Description) for every variable by looking them up in a sheet of the
    original workbook (e.g. old glossary) or a dictionary. The first occurance of the variable in the old sheet is used.

    Parameters
    ----------
    glossary_df : pd.DataFrame
        The dataframe that already contains the three columns
        ``["Sheet", "Variable_name", "Unit_or_type"]`` created by ``build_glossary_dataframe``.
    src_path : str
        Path to the original workbook that holds the reference “Glossary” sheet.
    glossary_sheet_name : str, default "Glossary"
        Name of the sheet that stores the master list of variables.
    var_col : str, default "Variable_name"
        Column in the reference sheet that holds the variable identifiers.
    col1 : str, default "Code_Query"
        Column of glossary_sheet_name that holds values to append to glossary_df.
    col2 : str, default "Description"
        Column of glossary_sheet_name that holds values to append to glossary_df.

    Returns
    -------
    pd.DataFrame
        ``glossary_df`` with two additional rows for each variable.  If a variable cannot be found, the added rows contain
        empty strings.
    """
    ref_glossary = pd.read_excel(
        src_path,
        sheet_name=glossary_sheet_name,
        dtype=str, skiprows=3        
    )
   
    ref_cols = ref_glossary [[var_col, col1, col2]]
    ref_cols_unique = (ref_cols.groupby(var_col, as_index=False).agg({col1: "first", col2: "first"})) #make a unique "dictionary" to avoid dublications
        
    enriched = pd.merge(glossary_df, ref_cols_unique, how="left", on = var_col)

    return enriched


if __name__ == "__main__":
    
    #provide the name of an input file that is located in the same folder as the script
    input_file = "ICASA_for_agroforstry_draft_4.xlsx"
    variables_all = "variable_sorting.xlsx"
    #provide a name of the output file
    output_file = "glossary.xlsx"
    
    BASE_DIR = os.path.abspath(os.path.dirname(__file__)) #do not run this line alone, only works when entire scrip is run
    input_path = os.path.join(BASE_DIR, input_file)
    output_path = os.path.join(BASE_DIR, output_file)
    
    print_sheet_names(input_path)
    glossary = build_glossary_dataframe(input_path, (2,3))
    enriched = enrich_glossary_with_metadata(glossary, input_file)
    write_glossary_to_new_file(enriched, output_path)
    